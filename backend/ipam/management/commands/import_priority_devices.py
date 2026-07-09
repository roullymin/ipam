import json
from collections import Counter
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import OperationalError, ProgrammingError, transaction
from django.utils import timezone
from openpyxl import load_workbook

from ipam.domains.config_backup.services import ConfigBackupConnectionError, ConfigBackupError, test_secret_login
from ipam.domains.vault.services import VaultError, read_secret, write_secret
from ipam.models import ConfigBackupTarget, IPAddress, RackDevice, SecretRecord


COLUMN_MAP = {
    'name': 0,
    'management_ip': 1,
    'device_type': 2,
    'vendor': 3,
    'ssh_port': 4,
    'username': 5,
    'password': 6,
    'backup_enabled': 7,
    'ansible_enabled': 8,
}


def _yes(value):
    return str(value or '').strip().lower() in {'yes', 'y', 'true', '1', '是', '启用', '纳入'}


def _backup_device_type(raw_type, name=''):
    text = f'{raw_type} {name}'.lower()
    if '防火' in text or 'firewall' in text or 'fw' in text:
        return 'firewall'
    if '核心' in text and ('交换' in text or 'switch' in text):
        return 'switch_core'
    if '接入' in text and ('交换' in text or 'switch' in text):
        return 'switch_access'
    if '交换' in text or 'switch' in text:
        return 'switch'
    if '路由' in text or 'router' in text:
        return 'router'
    if '安全' in text:
        return 'security'
    return 'other'


def _command_profile(vendor):
    text = str(vendor or '').lower()
    if 'h3c' in text:
        return 'h3c_comware'
    if 'cisco' in text:
        return 'cisco_ios'
    return 'huawei_vrp'


def _safe_int(value, default):
    try:
        return int(value if value not in (None, '') else default)
    except (TypeError, ValueError):
        return default


def _read_rows(path):
    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook.active
    rows = []
    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        values = ['' if value is None else str(value).strip() for value in row[:9]]
        if not any(values):
            continue
        rows.append(
            {
                'row': row_number,
                'name': values[COLUMN_MAP['name']],
                'management_ip': values[COLUMN_MAP['management_ip']],
                'device_type': values[COLUMN_MAP['device_type']],
                'vendor': values[COLUMN_MAP['vendor']],
                'ssh_port': _safe_int(values[COLUMN_MAP['ssh_port']], 22),
                'username': values[COLUMN_MAP['username']],
                'password': values[COLUMN_MAP['password']],
                'backup_enabled': _yes(values[COLUMN_MAP['backup_enabled']]),
                'ansible_enabled': _yes(values[COLUMN_MAP['ansible_enabled']]),
            }
        )
    return rows


def _match_asset(row):
    management_ip = row['management_ip']
    rack_device = None
    ip_asset = None
    try:
        if management_ip:
            rack_device = RackDevice.objects.select_related('rack', 'rack__datacenter').filter(mgmt_ip__iexact=management_ip).first()
            ip_asset = IPAddress.objects.filter(ip_address=management_ip).first()
        if rack_device is None and row['name']:
            rack_device = RackDevice.objects.select_related('rack', 'rack__datacenter').filter(name__iexact=row['name']).first()
        if ip_asset is None and row['name']:
            ip_asset = IPAddress.objects.filter(device_name__iexact=row['name']).first()
    except (OperationalError, ProgrammingError):
        return None, None
    return rack_device, ip_asset


def _credential_defaults(row, rack_device=None, ip_asset=None):
    target_type = 'general'
    if rack_device is not None:
        target_type = 'device'
    elif ip_asset is not None:
        target_type = 'ip'
    return {
        'name': f"{row['name'] or row['management_ip']} SSH 登录"[:160],
        'credential_type': 'ssh',
        'target_type': target_type,
        'rack_device': rack_device,
        'ip_address': ip_asset if rack_device is None else None,
        'datacenter': rack_device.rack.datacenter if rack_device and rack_device.rack_id and rack_device.rack.datacenter_id else None,
        'rack': rack_device.rack if rack_device and rack_device.rack_id else None,
        'username_hint': row['username'],
        'owner_team': '',
        'environment': 'production',
        'sensitivity': 'confidential',
        'rotation_days': 90,
        'status': 'active',
        'notes': f"重点设备清单导入：{row['name']} {row['management_ip']}".strip(),
    }


def _find_credential(row, rack_device=None, ip_asset=None):
    try:
        queryset = SecretRecord.objects.filter(username_hint=row['username'], status='active').order_by('-updated_at', '-id')
        if rack_device is not None:
            matched = queryset.filter(target_type='device', rack_device=rack_device).first()
            if matched is not None:
                return matched
        if ip_asset is not None:
            matched = queryset.filter(target_type='ip', ip_address=ip_asset).first()
            if matched is not None:
                return matched
        return queryset.filter(name__iexact=f"{row['name'] or row['management_ip']} SSH 登录"[:160]).first()
    except (OperationalError, ProgrammingError):
        return None


def _create_or_update_credential(row, rack_device=None, ip_asset=None):
    existing = _find_credential(row, rack_device, ip_asset)
    defaults = _credential_defaults(row, rack_device, ip_asset)
    if existing is None:
        credential = SecretRecord.objects.create(**defaults)
        action = 'created'
    else:
        for field, value in defaults.items():
            setattr(existing, field, value)
        existing.save()
        credential = existing
        action = 'updated'
    write_secret(
        credential.vault_path,
        row['username'],
        row['password'],
        {'record_id': credential.pk, 'name': credential.name, 'source': 'priority_device_import'},
    )
    credential.last_rotated_at = timezone.now()
    credential.save(update_fields=['last_rotated_at', 'updated_at'])
    return credential, action


def _provision_backup_target(row, credential, rack_device=None, ip_asset=None):
    target, created = ConfigBackupTarget.objects.update_or_create(
        management_ip=row['management_ip'],
        defaults={
            'name': row['name'] or row['management_ip'],
            'rack_device': rack_device,
            'ip_address': ip_asset if rack_device is None else None,
            'device_type': _backup_device_type(row['device_type'], row['name']),
            'command_profile': _command_profile(row['vendor']),
            'ssh_port': row['ssh_port'] or 22,
            'timeout_seconds': 30,
            'save_before_backup': True,
            'retention_count': 1,
            'credential': credential,
            'enabled': True,
        },
    )
    return target, 'created' if created else 'updated'


def _test_login(row, credential):
    try:
        payload = test_secret_login(
            credential=credential,
            management_ip=row['management_ip'],
            read_secret=read_secret,
            ssh_port=row['ssh_port'] or 22,
            timeout_seconds=30,
        )
        return 'success', payload.get('message') or 'SSH login succeeded.'
    except (ConfigBackupConnectionError, ConfigBackupError, VaultError) as exc:
        return 'failed', str(exc)


def _login_category(detail):
    text = str(detail or '').lower()
    if 'no acceptable kex' in text or 'ssh peer algorithms' in text or 'kex' in text or 'cipher' in text or 'hostkey' in text:
        return 'ssh_algorithm'
    if (
        'authentication' in text
        or 'auth' in text
        or 'password' in text
        or 'permission denied' in text
        or '认证' in text
        or '账号' in text
        or '账户' in text
        or '密码' in text
        or '登录' in text
    ):
        return 'auth_failed'
    if 'timed out' in text or 'timeout' in text:
        return 'timeout'
    if 'connection refused' in text:
        return 'refused'
    if 'no route to host' in text or 'network is unreachable' in text:
        return 'unreachable'
    if 'name or service not known' in text or 'temporary failure in name resolution' in text:
        return 'resolve_failed'
    return 'other'


class Command(BaseCommand):
    help = 'Validate/import a priority device Excel list and optionally test SSH login from the server network.'

    def add_arguments(self, parser):
        parser.add_argument('--file', required=True, help='Excel file path. Expected columns: device name, management IP, type, vendor, SSH port, username, password.')
        parser.add_argument('--write-secrets', action='store_true', help='Create/update SecretRecord rows and write passwords into OpenBao.')
        parser.add_argument('--provision-backup', action='store_true', help='Create/update ConfigBackupTarget rows for rows marked for backup.')
        parser.add_argument('--test-login', action='store_true', help='Run SSH login tests for rows that have usable credentials.')
        parser.add_argument('--output', default='', help='Optional JSON report path.')

    def handle(self, *args, **options):
        path = Path(options['file'])
        if not path.exists():
            raise CommandError(f'File not found: {path}')

        rows = _read_rows(path)
        duplicate_ips = [ip for ip, count in Counter(row['management_ip'] for row in rows if row['management_ip']).items() if count > 1]
        results = []
        counters = Counter()
        write_secrets = options['write_secrets']
        provision_backup = options['provision_backup']
        test_login = options['test_login']

        for row in rows:
            item = {
                'row': row['row'],
                'name': row['name'],
                'management_ip': row['management_ip'],
                'device_type': row['device_type'],
                'vendor': row['vendor'],
                'ssh_port': row['ssh_port'],
                'username': row['username'],
                'backup_enabled': row['backup_enabled'],
                'ansible_enabled': row['ansible_enabled'],
                'status': 'validated',
                'messages': [],
            }
            if not row['management_ip']:
                item['status'] = 'failed'
                item['messages'].append('missing management_ip')
                counters['failed'] += 1
                results.append(item)
                continue
            if not row['username']:
                item['status'] = 'failed'
                item['messages'].append('missing username')
                counters['failed'] += 1
                results.append(item)
                continue
            if not row['password']:
                item['status'] = 'failed'
                item['messages'].append('missing password')
                counters['failed'] += 1
                results.append(item)
                continue

            rack_device, ip_asset = _match_asset(row)
            item['asset_match'] = 'rack_device' if rack_device else ('ip_address' if ip_asset else 'unmatched')
            item['rack_device_id'] = rack_device.pk if rack_device else None
            item['ip_address_id'] = ip_asset.pk if ip_asset else None

            credential = _find_credential(row, rack_device, ip_asset)
            if write_secrets:
                try:
                    with transaction.atomic():
                        credential, action = _create_or_update_credential(row, rack_device, ip_asset)
                    item['credential_action'] = action
                    item['credential_id'] = credential.pk
                except VaultError as exc:
                    item['status'] = 'failed'
                    item['messages'].append(f'vault error: {exc}')
                    counters['failed'] += 1
                    results.append(item)
                    continue
            elif credential is not None:
                item['credential_id'] = credential.pk
                item['credential_action'] = 'existing'
            else:
                item['credential_action'] = 'dry_run'

            if provision_backup:
                if credential is None:
                    item['messages'].append('skip backup target: missing credential; rerun with --write-secrets first')
                else:
                    target, action = _provision_backup_target(row, credential, rack_device, ip_asset)
                    item['backup_target_action'] = action
                    item['backup_target_id'] = target.pk

            if test_login:
                if credential is None:
                    item['login_status'] = 'skipped'
                    item['login_detail'] = 'missing credential; rerun with --write-secrets first'
                    counters['login_skipped'] += 1
                else:
                    status, detail = _test_login(row, credential)
                    item['login_status'] = status
                    item['login_detail'] = detail
                    item['login_category'] = 'success' if status == 'success' else _login_category(detail)
                    counters[f'login_{status}'] += 1
                    if item['login_category'] != status:
                        counters[f"login_{item['login_category']}"] += 1

            if item['status'] == 'validated':
                counters['validated'] += 1
            results.append(item)

        report = {
            'source': str(path),
            'total': len(rows),
            'duplicate_ips': duplicate_ips,
            'summary': dict(counters),
            'results': results,
        }

        output = options.get('output')
        if not output:
            output_dir = Path(settings.BASE_DIR) / 'media' / 'reports'
            output_dir.mkdir(parents=True, exist_ok=True)
            output = output_dir / 'priority_device_import_report.json'
        output_path = Path(output)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding='utf-8')

        self.stdout.write(json.dumps({'total': len(rows), 'duplicate_ips': duplicate_ips, 'summary': dict(counters), 'report': str(output_path)}, ensure_ascii=False))
